import pandas as pd
import sys

file_path = r'C:\Users\KleinZebastianCatapa\Documents\INSIGHTEDCODES2026\public\Masterlist 2026-2030 139706 CL - with Cong-Gov-Mayor.xlsx'

try:
    # Read just the first few rows to get headers
    df = pd.read_excel(file_path, nrows=5)
    print("Columns:", df.columns.tolist())
    print("Types:", df.dtypes)
    print("Sample Data:")
    print(df.head())
except ImportError:
    print("pandas not installed, trying openpyxl directly...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = row
            break
        print("Headers from openpyxl:", headers)
        
        # Get first row of data
        data = []
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            data = row
            break
        print("First row data:", data)
    except ImportError:
        print("Neither pandas nor openpyxl is installed. Please install openpyxl: pip install openpyxl")
except Exception as e:
    print(f"Error: {e}")
